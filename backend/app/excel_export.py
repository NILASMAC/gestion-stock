import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

EXCEL_DIR = os.path.join(os.path.dirname(__file__), '..', 'exports')

def get_chemin_fichier_jour():
    os.makedirs(EXCEL_DIR, exist_ok=True)
    today = datetime.utcnow().strftime('%Y-%m-%d')
    return os.path.join(EXCEL_DIR, f'factures_{today}.xlsx')

def ajouter_vente_excel(vente, gerant, produit):
    chemin = get_chemin_fichier_jour()

    if os.path.exists(chemin):
        wb = load_workbook(chemin)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = f"Factures {datetime.utcnow().strftime('%d/%m/%Y')}"

        headers = ['N° Facture', 'Date', 'Produit', 'Gérant', 'Boutique',
                   'Client', 'Téléphone Client', 'Quantité', 'Prix Unitaire', 'Prix Total', 'Statut']
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='1E3A5F')
            cell.alignment = Alignment(horizontal='center')

        largeurs = [20, 18, 20, 20, 20, 20, 18, 10, 15, 15, 12]
        for i, l in enumerate(largeurs, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = l

    ws.append([
        vente.numero_facture,
        vente.date_vente.strftime('%d/%m/%Y %H:%M'),
        produit.nom if produit else '',
        gerant.nom if gerant else '',
        gerant.nom_boutique if gerant else '',
        vente.client_nom,
        vente.client_telephone or '',
        vente.quantite,
        vente.prix_unitaire,
        vente.prix_total,
        vente.statut,
    ])

    wb.save(chemin)